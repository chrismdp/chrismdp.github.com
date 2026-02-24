#!/usr/bin/env python3
"""
Generate certificate URLs for training course attendees.

Reads questionnaire data from an Excel spreadsheet, deduplicates by email,
derives full names from email addresses where needed, assigns courses based
on company/email domain, and generates certificate URLs.

Usage:
    python scripts/generate_certificate_urls.py [--input PATH] [--output PATH]
                                                [--genomics-date DATE]
                                                [--instem-date DATE]
"""

import argparse
import base64
import json
import re
import sys
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Configuration defaults
# ---------------------------------------------------------------------------

DEFAULT_INPUT = Path.home() / "vault" / "tmp" / "questionnaires.xlsx"
DEFAULT_OUTPUT = Path.home() / "vault" / "tmp" / "certificate-urls.tsv"
CERT_BASE_URL = "https://www.chrismdp.com/training/certificates/"

COURSE_DATES = {
    "AI Intensive": "July 2025",
    "Agentic Coding": "February 2026",
}

# Junk entries to filter out
JUNK_EMAILS = {"blank@g.com"}
JUNK_NAMES = {"."}


# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------

def titlecase_part(part: str) -> str:
    """Title-case a single name part, handling special cases.

    Handles:
    - Apostrophe names: o'connor -> O'Connor
    - Hyphenated names: jean-charles -> Jean-Charles, yonova-doing -> Yonova-Doing
    """
    # Handle hyphenated names: split, capitalise each segment, rejoin
    if "-" in part:
        return "-".join(titlecase_part(seg) for seg in part.split("-"))
    # Handle names with apostrophes: o'connor -> O'Connor
    if "'" in part:
        segments = part.split("'")
        return "'".join(s.capitalize() for s in segments)
    return part.capitalize()


def derive_name_from_email(email: str) -> str | None:
    """
    Attempt to derive a full name from the local part of an email address.

    Splits on '.' and title-cases each part.  Returns None if the local part
    has no '.' (i.e. we can't reliably split first/last) or looks like a
    username rather than a real name.
    """
    local = email.split("@")[0]
    # Strip trailing digits (e.g. thomas.hickman42 -> thomas.hickman)
    local = re.sub(r"\d+$", "", local)
    parts = local.split(".")
    if len(parts) < 2:
        # Single-word local part -- can't derive first + last reliably
        return None
    return " ".join(titlecase_part(p) for p in parts if p)


def normalise_name(first_name: str | None, last_name: str | None,
                   email: str) -> str | None:
    """
    Get the first name from form fields. Returns None if no name provided.

    Certificates use first names only. No name in the form = no certificate.
    """
    first = (first_name or "").strip()

    # Filter junk names
    if first in JUNK_NAMES:
        return None

    if not first:
        return None

    # If it's a full name like "Guillermo Reales", take just the first name
    parts = first.split()
    return titlecase_part(parts[0])


# ---------------------------------------------------------------------------
# Course assignment
# ---------------------------------------------------------------------------

def assign_course(email: str, company: str | None) -> str:
    """Assign a course based on company name and email domain."""
    email_lower = email.lower()
    company_lower = (company or "").lower()

    if "genomics" in company_lower or "genomics" in email_lower:
        return "AI Intensive"
    if "instem" in company_lower or "instem" in email_lower:
        return "Agentic Coding"
    return "UNKNOWN"


# ---------------------------------------------------------------------------
# Spreadsheet reading
# ---------------------------------------------------------------------------

def read_sheet(ws, sheet_name: str) -> list[dict]:
    """
    Read a worksheet and return a list of person dicts.

    Each dict has keys: first_name, last_name, email, company.
    Handles the different column naming conventions across sheets.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Build column index map
    col = {}
    for i, h in enumerate(headers):
        if h == "first_name":
            col["first_name"] = i
        elif h in ("email_address", "email"):
            col["email"] = i
        elif h == "company":
            col["company"] = i
        elif h == "last name":
            col["last_name"] = i

    people = []
    for row in rows[1:]:
        email_val = row[col["email"]] if "email" in col and col["email"] < len(row) else None
        if not email_val or not str(email_val).strip():
            continue  # skip rows with no email

        email = str(email_val).strip().lower()

        first = None
        if "first_name" in col and col["first_name"] < len(row):
            val = row[col["first_name"]]
            if val:
                first = str(val).strip()

        last = None
        if "last_name" in col and col["last_name"] < len(row):
            val = row[col["last_name"]]
            if val:
                last = str(val).strip()

        company = None
        if "company" in col and col["company"] < len(row):
            val = row[col["company"]]
            if val:
                company = str(val).strip()

        people.append({
            "first_name": first,
            "last_name": last,
            "email": email,
            "company": company,
            "source_sheet": sheet_name,
        })

    return people


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def dedupe_people(all_people: list[dict]) -> dict[str, dict]:
    """
    Deduplicate by email address, keeping the richest record.

    When merging, prefer non-None values. For names, prefer a record that
    has a fuller name (e.g. full name over just first name).
    """
    by_email: dict[str, dict] = {}

    for person in all_people:
        email = person["email"]

        # Skip junk
        if email in JUNK_EMAILS:
            continue
        first = (person.get("first_name") or "").strip()
        if first in JUNK_NAMES:
            continue

        if email not in by_email:
            by_email[email] = dict(person)
            continue

        existing = by_email[email]

        # Merge: prefer non-None / fuller values
        for key in ("first_name", "last_name", "company"):
            new_val = (person.get(key) or "").strip()
            old_val = (existing.get(key) or "").strip()
            if not old_val and new_val:
                existing[key] = new_val
            elif old_val and new_val:
                # Prefer the longer / more complete value
                if key == "first_name" and " " in new_val and " " not in old_val:
                    existing[key] = new_val

    return by_email


# ---------------------------------------------------------------------------
# Certificate URL generation
# ---------------------------------------------------------------------------

def generate_cert_url(name: str, email: str, course: str, date: str) -> str:
    """Generate a certificate URL with base64-encoded JSON payload."""
    payload = json.dumps({
        "name": name,
        "email": email,
        "course": course,
        "date": date,
    }, separators=(",", ":"))
    encoded = base64.b64encode(payload.encode("utf-8")).decode("ascii")
    return f"{CERT_BASE_URL}#{encoded}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate certificate URLs for training attendees."
    )
    parser.add_argument(
        "--input", "-i",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Path to questionnaires XLSX (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output TSV path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--genomics-date",
        default=COURSE_DATES["AI Intensive"],
        help=f"Date string for AI Intensive certificates (default: {COURSE_DATES['AI Intensive']})",
    )
    parser.add_argument(
        "--instem-date",
        default=COURSE_DATES["Agentic Coding"],
        help=f"Date string for Agentic Coding certificates (default: {COURSE_DATES['Agentic Coding']})",
    )
    args = parser.parse_args()

    course_dates = {
        "AI Intensive": args.genomics_date,
        "Agentic Coding": args.instem_date,
    }

    # Read all sheets
    wb = openpyxl.load_workbook(args.input, read_only=True)
    all_people = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_people.extend(read_sheet(ws, sheet_name))
    wb.close()

    # Deduplicate
    by_email = dedupe_people(all_people)

    # Build output rows
    rows = []
    unknowns = []
    skipped_no_name = []
    for email, person in sorted(by_email.items()):
        name = normalise_name(
            person.get("first_name"),
            person.get("last_name"),
            email,
        )

        # No name in form = no certificate
        if name is None:
            skipped_no_name.append(email)
            continue

        course = assign_course(email, person.get("company"))
        date = course_dates.get(course, "")

        if course == "UNKNOWN":
            unknowns.append((name, email, course, ""))
            continue

        url = generate_cert_url(name, email, course, date)
        rows.append((name, email, course, url))

    # Print warnings
    if skipped_no_name:
        print(f"SKIPPED (no first name): {len(skipped_no_name)} people", file=sys.stderr)
        for email in skipped_no_name:
            print(f"  {email}", file=sys.stderr)
        print(file=sys.stderr)

    if unknowns:
        print("WARNING: Could not assign course for:", file=sys.stderr)
        for name, email, _, _ in unknowns:
            print(f"  {name} <{email}>", file=sys.stderr)
        print(file=sys.stderr)

    # Output TSV
    header = "name\temail\tcourse\turl"
    lines = [header]
    for name, email, course, url in rows:
        lines.append(f"{name}\t{email}\t{course}\t{url}")

    # Also include unknowns at the end (flagged)
    for name, email, course, url in unknowns:
        lines.append(f"{name}\t{email}\t{course}\t")

    output = "\n".join(lines) + "\n"

    # Print to stdout
    sys.stdout.write(output)

    # Save to file
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(output, encoding="utf-8")
    print(f"\nSaved to {args.output}", file=sys.stderr)
    print(f"Total: {len(rows)} certificates, {len(skipped_no_name)} skipped (no name), {len(unknowns)} unknown", file=sys.stderr)


if __name__ == "__main__":
    main()
